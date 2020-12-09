import json
import openpyxl
import requests

# import logging
#
# logging.debug('This is debug message')
# logging.info('This is info message')
# logging.warning('This is warning message')

# host = 'http://192.168.2.252:18080'
# path ='/4000/4000/0/guest.medicine.getSearchPageData'
#
# headers = {
#   'Content-Type': 'application/x-www-form-urlencoded'
# }
#
#
# params={'keywords':'克拉霉素'}
# json=params
#
# response=requests.get(url=host+path,headers=headers,params=json)
# print(response.text)
# print(response.url)


class ExcelData:
    def openexl(self,path):
        wb=openpyxl.load_workbook(path)

        sheet=wb['Sheet1']
        # print(sheet.max_row)
        # print(sheet.max_column)

        data_list=[]
        for i in range(2,sheet.max_row+1):
            a=[]
            b=sheet.cell(i,1).value
            c=sheet.cell(i,2).value
            d=sheet.cell(i,3).value
            e=sheet.cell(i,4).value
            f=sheet.cell(i,5).value
            g = sheet.cell(i,6).value
            a.append(b)
            a.append(c)
            a.append(d)
            a.append(e)
            a.append(f)
            a.append(g)
            a.append(i)

            data_list.append(a)
        return data_list

# wb=openpyxl.load_workbook("E:\pythonbijia\data\case.xlsx")
# sheet=wb['Sheet1']
# print(sheet.max_row)
# print(sheet.max_column)
#
# data_list=[]
# for i in range(2,sheet.max_row+1):
#     a=[]
#     b=sheet.cell(i,1).value
#     c=sheet.cell(i,2).value
#     d=sheet.cell(i,3).value
#     e=sheet.cell(i,4).value
#     f=sheet.cell(i,5).value
#
#     a.append(b)
#     a.append(c)
#     a.append(d)
#     a.append(e)
#     a.append(f)
#
#     data_list.append(a)
# print(data_list)


