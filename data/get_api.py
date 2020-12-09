import json
import openpyxl


class ExcelData:
    def openexl(self,path,wksheet):
        wb=openpyxl.load_workbook(path)

        sheet = wb[wksheet]
        data_list = []
        for i in range(2,sheet.max_row+1):
            a=[]
            b=sheet.cell(i,1).value
            c=sheet.cell(i,2).value
            d=sheet.cell(i,3).value
            e=sheet.cell(i,4).value

            a.append(b)
            a.append(c)
            a.append(d)
            a.append(e)
            a.append(i)

            data_list.append(a)
        return data_list


