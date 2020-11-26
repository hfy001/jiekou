import json
import openpyxl

class excelshuju2:
    def openexl(self,lujing,wksheet):
        wb=openpyxl.load_workbook(lujing)

        sheet=wb[wksheet]
        # print(sheet.max_row)
        # print(sheet.max_column)

        data_list=[]

        for i in range(2,sheet.max_row+1):
            a=[]
            b=sheet.cell(i,1).value
            c=sheet.cell(i,2).value
            d=sheet.cell(i,3).value
            e=sheet.cell(i,4).value
            f=sheet.cell(i,5).value
            g = sheet.cell(i,6).value

            ''' 相对canshu1.py去掉了headers'''

            a.append(b)
            a.append(c)
            a.append(d)
            a.append(e)
            a.append(f)
            a.append(i)
            a.append(g)



            data_list.append(a)
        return data_list


