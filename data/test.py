import json
import openpyxl
import requests

# host = 'http://192.168.2.252:18080'
# path ='/4000/4000/0/guest.medicine.getShopMedicines'
# headers = {
#   'Content-Type': 'application/x-www-form-urlencoded'
# }
#
# conditions={
#         "sort":"",
#         "sorttype":"desc",
#         "medicineid":"278803",
#         "is_freepostage":"1",
#         "lat":"31.211457",
#         "lng":"121.635224",
#         "user_city_name":"上海市",
#         "user_region_id":"95"
#     }
# params={
#     'user_city_name':'上海市',
#     'user_region_id':'95',
#     'pageIndex':'1',
#     'pageSize':'10',
#     'conditions':json.dumps(conditions)
#     }


wb=openpyxl.load_workbook('canshu1.xlsx')
sheet=wb['Sheet1']
print(sheet.max_row)
print(sheet.max_column)


# def huoqushuju():
#     data_list=[]
#     for i in range(2,sheet.max_row+1):
#         a=[]
#         b=sheet.cell(i,1).value
#         c=sheet.cell(i,2).value
#         d=sheet.cell(i,3).value
#         e=sheet.cell(i,4).value
#         f=sheet.cell(i,5).value
#         a.append(b)
#         a.append(c)
#         a.append(d)
#         a.append(e)
#         a.append(f)
#         data_list.append(a)
#     return data_list
#
# data = huoqushuju()
# params = json.loads(data[0][4])
# headers = json.loads(data[0][2])
# print(params)
#
# response = requests.get(url=data[0][0]+data[0][1],headers=headers,params=params)
# print(response.content)
# print(response.url)

for l in range(2, sheet.max_row + 1):
    la = []
    lb = sheet.cell(l, 1).value
    lc = sheet.cell(l, 2).value
    ld = sheet.cell(l, 3).value
    le = sheet.cell(l, 4).value

    lf = sheet.cell(l, 5).value


    print(le)
